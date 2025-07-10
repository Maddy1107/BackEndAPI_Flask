# excel_functions.py
import datetime
from io import BytesIO
from openpyxl import load_workbook
import pandas as pd


def extract_product_quantity(file_path):
    try:
        df = pd.read_excel(file_path, header=None, skiprows=4)
        if len(df.columns) < 4:
            return {}, "Excel file does not have enough columns"
        product_col, quantity_col = df.columns[2], df.columns[3]

        df[quantity_col] = df[quantity_col].fillna("")

        filtered_df = df[[product_col, quantity_col]]
        data_dict = {
            str(row[product_col]): row[quantity_col]
            for _, row in filtered_df.iterrows()
            if pd.notnull(row[product_col]) and str(row[product_col]).strip() != ""
        }
        return data_dict, None
    except Exception as e:
        return {}, str(e)


def update_excel_file(file, update_dict, name, sheet_name=None):
    try:
        wb = load_workbook(file)
        sheet = wb[sheet_name] if sheet_name else wb.active

        month_col_index = get_month_column(sheet)

        for row in sheet.iter_rows(min_row=5):
            product_cell = row[2]
            target_cell = row[3] if month_col_index == -1 else row[month_col_index]
            remarks_cell = row[4] if month_col_index == -1 else row[month_col_index + 1]

            product_name = str(product_cell.value).strip() if product_cell.value else ""
            if product_name in update_dict:
                values = update_dict[product_name]
                if isinstance(values, list) and len(values) >= 2:
                    target_cell.value = values[0]
                    remarks_cell.value = values[1]
                else:
                    return (
                        None,
                        f"Invalid format for '{product_name}'. Expected a list of at least 2 values.",
                    )

        update_named_cells(sheet, name)

        output_stream = BytesIO()
        wb.save(output_stream)
        output_stream.seek(0)

        return output_stream, None
    except Exception as e:
        return None, str(e)


def get_month_column(sheet):
    current_month = datetime.datetime.now().strftime("%B").lower()  # e.g., "july"
    try:
        header_row = list(sheet.iter_rows(min_row=4, max_row=4))[
            0
        ]  # 5th row is index 4
    except IndexError:
        return -1  # Row doesn't exist

    for col_index, cell in enumerate(header_row):
        cell_value = str(cell.value).strip().lower() if cell.value else ""
        if current_month in cell_value:
            return col_index  # 0-based index

    return -1  # Not found


def update_named_cells(sheet, name):
    month_col = get_month_column(sheet)

    for row_idx in range(1, 5):  # Rows 1 to 4 (1-based)
        row = list(sheet.iter_rows(min_row=row_idx, max_row=row_idx))[0]

        for i, cell in enumerate(row):
            if not cell.value:
                continue

            val = str(cell.value).strip().lower()

            if "name" in val or "handover" in val:
                should_update = (month_col == -1) or (i == month_col)
                if should_update and i + 1 < len(row):
                    row[i + 1].value = (
                        name
                        if "name" in val
                        else datetime.datetime.now().strftime("%d-%m-%Y")
                    )
